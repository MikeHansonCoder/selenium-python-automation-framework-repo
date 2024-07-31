import inspect
import logging
from openpyxl import Workbook, load_workbook
import softest
import csv


class Utils(softest.TestCase):
    def asserting_elements(self, listofStops, value):
        for stops in listofStops:
            print("The text is: "+stops.text)
            self.soft_assert(self.assertEqual, stops.text, value)
            if stops.text == value:
                print("Assert Pass")
            else:
                print("Assert Fail")
        self.assert_all()

    def custom_logger(loggingLevel = logging.DEBUG):
        # Set class/method name from where its called
        logger_name = inspect.stack()[1][3]

        # Create Logger
        logger = logging.getLogger(logger_name)
        logger.setLevel(loggingLevel)

        #create console handeler or file handler and set the log level
        fh = logging.FileHandler("automation.log")

        #Create a formattor: How you want your logs to be formatted
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')

        #Add formattor to console or file handler
        fh.setFormatter(formatter)

        #Add handler to logger
        logger.addHandler(fh)
        return logger

    def read_data_from_exel(file_name, sheet):
        datalist = []

        wb = load_workbook(filename = file_name)
        sh = wb[sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column

        for i in range(2, row_ct+1):
            row = []
            for j in range(1, col_ct+1):
                row.append(sh.cell(row=i, column=j).value)
            datalist.append(row)
        return datalist

    def read_data_from_csv(file_name):
        #Create empty list
        datalist = []

        #Open CSV file
        csvdata = open(file_name, "r")

        #Create CSV reader
        reader = csv.reader(csvdata)

        #skip the header
        next(reader)

        #Add CSV rows in list
        for rows in reader:
            datalist.append(rows)
        return datalist


